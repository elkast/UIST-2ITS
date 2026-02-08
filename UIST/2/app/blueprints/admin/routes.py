"""
Routes d'administration pour UIST-Planify
Gestion CRUD des salles, enseignants, cours, filières
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from app.models import Salle, Enseignant, Utilisateur, Cours, Filiere, EmploiDuTemps, Etudiant, Presence, Note, Conflit, Parent
from app.utils import connexion_requise, role_required, generer_matricule
from datetime import datetime, timedelta
from flask import Response, session
import openpyxl
import io
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

admin_bp = Blueprint('admin', __name__)

@admin_bp.route('/test-models')
def test_models():
    """Route de test pour vérifier les imports"""
    try:
        from app.models_extended import Notification, Bulletin, ImportNote
        return f"✅ Import OK - Notification: {Notification}, Bulletin: {Bulletin}, ImportNote: {ImportNote}"
    except Exception as e:
        return f"❌ Erreur import: {str(e)}"

@admin_bp.route('/directeur-dashboard')
@role_required(['directeur', 'DIRECTEUR', 'ADMIN', 'SUPER_ADMIN'])
def directeur_dashboard():
    """
    Dashboard du Directeur - Validation des notes en temps réel
    """
    from app.models import Note, Message, Etudiant, Filiere
    
    # Statistiques
    notes_validees = Note.obtenir_notes_validees() or []
    notes_en_attente = Note.obtenir_notes_en_attente() or []
    
    # Compter les signalements (messages de type SIGNALEMENT)
    signalements = Message.obtenir_messages_non_lus(session['utilisateur_id']) or []
    signalements_count = len([m for m in signalements if m.get('type_message') == 'SIGNALEMENT'])
    
    # Nombre d'étudiants
    etudiants = Etudiant.obtenir_tous() or []
    
    stats = {
        'notes_validees': len(notes_validees),
        'notes_en_attente': len(notes_en_attente),
        'signalements': signalements_count,
        'nb_etudiants': len(etudiants)
    }
    
    # Liste des filières pour les filtres
    filieres = Filiere.obtenir_toutes() or []
    
    return render_template('directeur/dashboard.html', 
                         stats=stats, 
                         filieres=filieres)


@admin_bp.route('/gestionnaire-pv-dashboard')
@role_required(['GESTIONNAIRE_PV', 'ADMIN', 'SUPER_ADMIN'])
def gestionnaire_pv_dashboard():
    """
    Dashboard du Gestionnaire PV - Génération de bulletins
    """
    from app.models import Etudiant, Filiere
    try:
        from app.models_extended import Bulletin
    except ImportError:
        Bulletin = None
    
    # Statistiques
    bulletins = Bulletin.obtenir_tous() if Bulletin else []
    etudiants = Etudiant.obtenir_tous() or []
    filieres = Filiere.obtenir_toutes() or []
    
    stats = {
        'bulletins_generes': len(bulletins) if bulletins else 0,
        'nb_etudiants': len(etudiants),
        'nb_filieres': len(filieres),
        'taux_reussite': 0  # À calculer
    }
    
    # Stats par filière
    stats_filieres = []
    
    return render_template('gest_pv/dashboard_enhanced.html',
                         stats=stats,
                         bulletins=bulletins or [],
                         filieres=filieres,
                         stats_filieres=stats_filieres,
                         filiere_id=None,
                         semestre=None,
                         annee=None)


@admin_bp.route('/gestionnaire-examens-dashboard')
@role_required(['GESTIONNAIRE_EXAMENS', 'ADMIN', 'SUPER_ADMIN'])
def gestionnaire_examens_dashboard():
    """
    Dashboard du Gestionnaire Examens - Structuration et import
    """
    from app.models import Cours, Filiere
    try:
        from app.models_extended import ImportNote
    except ImportError:
        ImportNote = None
    
    # Statistiques
    imports = ImportNote.obtenir_historique() if ImportNote else []
    cours = Cours.obtenir_tous() or []
    filieres = Filiere.obtenir_toutes() or []
    
    stats = {
        'imports_total': len(imports) if imports else 0,
        'imports_succes': 0,
        'imports_erreurs': 0,
        'nb_cours': len(cours),
        'nb_filieres': len(filieres),
        'notes_total': 0
    }
    
    return render_template('gest_exam/dashboard_enhanced.html',
                         stats=stats,
                         imports=imports or [],
                         cours=cours,
                         filieres=filieres)


@admin_bp.route('/tableau-bord')
@role_required(['administration', 'sous_admin', 'directeur', 'ADMIN', 'SUPER_ADMIN'])
def tableau_bord():
    """
    Tableau de bord de l'administration
    """
    # Statistiques
    salles = Salle.obtenir_toutes()
    enseignants = Enseignant.obtenir_tous()
    cours = Cours.obtenir_tous()
    filieres = Filiere.obtenir_toutes()
    creneaux = EmploiDuTemps.obtenir_tous()
    etudiants = Etudiant.obtenir_tous()

    stats = {
        'nb_salles': len(salles) if salles else 0,
        'nb_enseignants': len(enseignants) if enseignants else 0,
        'nb_cours': len(cours) if cours else 0,
        'nb_filieres': len(filieres) if filieres else 0,
        'nb_creneaux': len(creneaux) if creneaux else 0,
        'nb_etudiants': len(etudiants) if etudiants else 0
    }

    # Statistiques supplémentaires
    from app.models import Note, Presence, Utilisateur
    notes_total = Note.obtenir_toutes()
    presences_total = Presence.obtenir_toutes_presences()

    stats.update({
        'nb_notes': len(notes_total) if notes_total else 0,
        'nb_presences': len(presences_total) if presences_total else 0,
        'nb_utilisateurs': len(Utilisateur.obtenir_tous()) if Utilisateur.obtenir_tous() else 0
    })

    # Récupérer les conflits actifs
    conflits = Conflit.obtenir_conflits_actifs()

    return render_template('admin/dashboard.html', stats=stats, conflits=conflits)

# ==================== GESTION DES SALLES ====================

@admin_bp.route('/salles')
@role_required(['administration', 'directeur'])
def gestion_salles():
    """
    Page de gestion des salles
    """
    salles = Salle.obtenir_toutes()
    return render_template('admin/gestion_salles.html', salles=salles)

@admin_bp.route('/salles/ajouter', methods=['POST'])
@role_required(['administration', 'directeur'])
def ajouter_salle():
    """
    Ajouter une nouvelle salle
    """
    nom_salle = request.form.get('nom_salle', '').strip()
    capacite = request.form.get('capacite', '').strip()
    equipements = request.form.get('equipements', '').strip()
    
    # Validation
    if not nom_salle or not capacite:
        flash('Le nom et la capacité sont obligatoires.', 'danger')
        return redirect(url_for('admin.gestion_salles'))
    
    try:
        capacite = int(capacite)
        if capacite <= 0:
            raise ValueError
    except ValueError:
        flash('La capacité doit être un nombre positif.', 'danger')
        return redirect(url_for('admin.gestion_salles'))
    
    # Créer la salle
    salle_id = Salle.creer(nom_salle, capacite, equipements)
    
    if salle_id:
        flash(f'Salle "{nom_salle}" ajoutée avec succès.', 'success')
    else:
        flash('Erreur lors de l\'ajout de la salle.', 'danger')
    
    return redirect(url_for('admin.gestion_salles'))

@admin_bp.route('/salles/modifier/<int:salle_id>', methods=['POST'])
@role_required(['administration', 'directeur'])
def modifier_salle(salle_id):
    """
    Modifier une salle existante
    """
    nom_salle = request.form.get('nom_salle', '').strip()
    capacite = request.form.get('capacite', '').strip()
    equipements = request.form.get('equipements', '').strip()
    
    # Validation
    if not nom_salle or not capacite:
        flash('Le nom et la capacité sont obligatoires.', 'danger')
        return redirect(url_for('admin.gestion_salles'))
    
    try:
        capacite = int(capacite)
        if capacite <= 0:
            raise ValueError
    except ValueError:
        flash('La capacité doit être un nombre positif.', 'danger')
        return redirect(url_for('admin.gestion_salles'))
    
    # Modifier la salle
    resultat = Salle.modifier(salle_id, nom_salle, capacite, equipements)
    
    if resultat:
        flash(f'Salle "{nom_salle}" modifiée avec succès.', 'success')
    else:
        flash('Erreur lors de la modification de la salle.', 'danger')
    
    return redirect(url_for('admin.gestion_salles'))

@admin_bp.route('/salles/supprimer/<int:salle_id>', methods=['POST'])
@role_required(['administration', 'directeur'])
def supprimer_salle(salle_id):
    """
    Supprimer une salle
    """
    resultat = Salle.supprimer(salle_id)
    
    if resultat:
        flash('Salle supprimée avec succès.', 'success')
    else:
        flash('Erreur lors de la suppression de la salle.', 'danger')
    
    return redirect(url_for('admin.gestion_salles'))

# ==================== GESTION DES FILIÈRES ====================

@admin_bp.route('/filieres')
@role_required(['administration', 'directeur'])
def gestion_filieres():
    """
    Page de gestion des filières
    """
    filieres = Filiere.obtenir_toutes()
    return render_template('admin/academic_struct.html', filieres=filieres)

@admin_bp.route('/filieres/ajouter', methods=['POST'])
@role_required(['administration', 'directeur'])
def ajouter_filiere():
    """
    Ajouter une nouvelle filière
    """
    nom_filiere = request.form.get('nom_filiere', '').strip()
    niveau = request.form.get('niveau', '').strip()
    nombre_etudiants = request.form.get('nombre_etudiants', '0').strip()
    
    # Validation
    if not nom_filiere or not niveau:
        flash('Le nom et le niveau sont obligatoires.', 'danger')
        return redirect(url_for('admin.gestion_filieres'))
    
    try:
        nombre_etudiants = int(nombre_etudiants)
        if nombre_etudiants < 0:
            raise ValueError
    except ValueError:
        flash('Le nombre d\'étudiants doit être un nombre positif ou zéro.', 'danger')
        return redirect(url_for('admin.gestion_filieres'))
    
    # Créer la filière
    filiere_id = Filiere.creer(nom_filiere, niveau, nombre_etudiants)
    
    if filiere_id:
        flash(f'Filière "{nom_filiere}" ajoutée avec succès.', 'success')
    else:
        flash('Erreur lors de l\'ajout de la filière.', 'danger')
    
    return redirect(url_for('admin.gestion_filieres'))

@admin_bp.route('/filieres/modifier/<int:filiere_id>', methods=['POST'])
@role_required(['administration', 'directeur'])
def modifier_filiere(filiere_id):
    """
    Modifier une filière existante
    """
    nom_filiere = request.form.get('nom_filiere', '').strip()
    niveau = request.form.get('niveau', '').strip()
    nombre_etudiants = request.form.get('nombre_etudiants', '0').strip()
    
    # Validation
    if not nom_filiere or not niveau:
        flash('Le nom et le niveau sont obligatoires.', 'danger')
        return redirect(url_for('admin.gestion_filieres'))
    
    try:
        nombre_etudiants = int(nombre_etudiants)
        if nombre_etudiants < 0:
            raise ValueError
    except ValueError:
        flash('Le nombre d\'étudiants doit être un nombre positif ou zéro.', 'danger')
        return redirect(url_for('admin.gestion_filieres'))
    
    # Modifier la filière
    resultat = Filiere.modifier(filiere_id, nom_filiere, niveau, nombre_etudiants)
    
    if resultat:
        flash(f'Filière "{nom_filiere}" modifiée avec succès.', 'success')
    else:
        flash('Erreur lors de la modification de la filière.', 'danger')
    
    return redirect(url_for('admin.gestion_filieres'))

@admin_bp.route('/filieres/supprimer/<int:filiere_id>', methods=['POST'])
@role_required(['administration', 'directeur'])
def supprimer_filiere(filiere_id):
    """
    Supprimer une filière
    """
    resultat = Filiere.supprimer(filiere_id)
    
    if resultat:
        flash('Filière supprimée avec succès.', 'success')
    else:
        flash('Erreur lors de la suppression de la filière.', 'danger')
    
    return redirect(url_for('admin.gestion_filieres'))

# ==================== GESTION DES ENSEIGNANTS ====================

@admin_bp.route('/enseignants')
@role_required(['administration', 'directeur'])
def gestion_enseignants():
    """
    Page de gestion des enseignants
    """
    enseignants = Enseignant.obtenir_tous()
    return render_template('admin/gestion_enseignants.html', enseignants=enseignants)

@admin_bp.route('/enseignants/ajouter', methods=['POST'])
@role_required(['administration'])
def ajouter_enseignant():
    """
    Ajouter un nouvel enseignant (Redirige vers gestion_utilisateurs)
    """
    flash('Veuillez utiliser le module "Gestion des Utilisateurs" pour ajouter un enseignant.', 'info')
    return redirect(url_for('admin.gestion_utilisateurs'))

@admin_bp.route('/enseignants/modifier/<int:enseignant_id>', methods=['POST'])
@role_required(['administration', 'directeur'])
def modifier_enseignant(enseignant_id):
    """
    Modifier un enseignant existant
    """
    specialite = request.form.get('specialite', '').strip()
    
    # Modifier la spécialité
    resultat = Enseignant.modifier(enseignant_id, specialite)
    
    if resultat:
        flash('Enseignant modifié avec succès.', 'success')
    else:
        flash('Erreur lors de la modification de l\'enseignant.', 'danger')
    
    return redirect(url_for('admin.gestion_enseignants'))

@admin_bp.route('/enseignants/supprimer/<int:enseignant_id>', methods=['POST'])
@role_required(['administration', 'directeur'])
def supprimer_enseignant(enseignant_id):
    """
    Supprimer un enseignant
    """
    resultat = Utilisateur.supprimer(enseignant_id)
    
    if resultat:
        flash('Enseignant supprimé avec succès.', 'success')
    else:
        flash('Erreur lors de la suppression de l\'enseignant.', 'danger')
    
    return redirect(url_for('admin.gestion_enseignants'))

# ==================== GESTION DES COURS ====================

@admin_bp.route('/cours')
@role_required(['administration', 'directeur'])
def gestion_cours():
    """
    Page de gestion des cours
    """
    cours = Cours.obtenir_tous()
    filieres = Filiere.obtenir_toutes()
    enseignants = Enseignant.obtenir_tous()
    salles = Salle.obtenir_toutes()
    
    return render_template('admin/gestion_cours.html', 
                         cours=cours, 
                         filieres=filieres,
                         enseignants=enseignants,
                         salles=salles)

@admin_bp.route('/cours/ajouter', methods=['POST'])
@role_required(['administration', 'directeur'])
def ajouter_cours():
    """
    Ajouter un nouveau cours
    """
    nom_cours = request.form.get('nom_cours', '').strip()
    filiere_id = request.form.get('filiere_id', '').strip()
    type_cours = request.form.get('type_cours', '').strip()
    
    # Validation
    if not all([nom_cours, filiere_id, type_cours]):
        flash('Tous les champs sont obligatoires.', 'danger')
        return redirect(url_for('admin.gestion_cours'))
    
    try:
        filiere_id = int(filiere_id)
    except ValueError:
        flash('Filière invalide.', 'danger')
        return redirect(url_for('admin.gestion_cours'))
    
    # Créer le cours
    cours_id = Cours.creer(nom_cours, filiere_id, type_cours)
    
    if cours_id:
        flash(f'Cours "{nom_cours}" ajouté avec succès.', 'success')
    else:
        flash('Erreur lors de l\'ajout du cours.', 'danger')
    
    return redirect(url_for('admin.gestion_cours'))

@admin_bp.route('/cours/modifier/<int:cours_id>', methods=['POST'])
@role_required(['administration', 'directeur'])
def modifier_cours(cours_id):
    """
    Modifier un cours existant
    """
    nom_cours = request.form.get('nom_cours', '').strip()
    filiere_id = request.form.get('filiere_id', '').strip()
    type_cours = request.form.get('type_cours', '').strip()
    
    # Validation
    if not all([nom_cours, filiere_id, type_cours]):
        flash('Tous les champs sont obligatoires.', 'danger')
        return redirect(url_for('admin.gestion_cours'))
    
    try:
        filiere_id = int(filiere_id)
    except ValueError:
        flash('Filière invalide.', 'danger')
        return redirect(url_for('admin.gestion_cours'))
    
    # Modifier le cours
    resultat = Cours.modifier(cours_id, nom_cours, filiere_id, type_cours)
    
    if resultat:
        flash(f'Cours "{nom_cours}" modifié avec succès.', 'success')
    else:
        flash('Erreur lors de la modification du cours.', 'danger')
    
    return redirect(url_for('admin.gestion_cours'))

@admin_bp.route('/cours/supprimer/<int:cours_id>', methods=['POST'])
@role_required(['administration', 'directeur'])
def supprimer_cours(cours_id):
    """
    Supprimer un cours
    """
    resultat = Cours.supprimer(cours_id)
    
    if resultat:
        flash('Cours supprimé avec succès.', 'success')
    else:
        flash('Erreur lors de la suppression du cours.', 'danger')
    
    return redirect(url_for('admin.gestion_cours'))

@admin_bp.route('/cours/assigner-enseignant/<int:cours_id>', methods=['POST'])
@role_required(['administration', 'directeur', 'ADMIN', 'SUPER_ADMIN'])
def assigner_enseignant_cours(cours_id):
    """
    Assigner un enseignant à un cours via un créneau EDT
    """
    enseignant_id = request.form.get('enseignant_id', '').strip()
    salle_id = request.form.get('salle_id', '').strip()
    jour = request.form.get('jour', '').strip()
    heure_debut = request.form.get('heure_debut', '').strip()
    heure_fin = request.form.get('heure_fin', '').strip()
    
    # Validation
    if not all([enseignant_id, salle_id, jour, heure_debut, heure_fin]):
        flash('Tous les champs sont obligatoires pour assigner un enseignant.', 'danger')
        return redirect(url_for('admin.gestion_cours'))
    
    try:
        enseignant_id = int(enseignant_id)
        salle_id = int(salle_id)
    except ValueError:
        flash('Données invalides.', 'danger')
        return redirect(url_for('admin.gestion_cours'))
    
    # Vérifier les conflits
    conflit = EmploiDuTemps.verifier_conflit(enseignant_id, salle_id, jour, heure_debut, heure_fin, cours_id)
    
    if conflit:
        type_conflit = conflit['type']
        if type_conflit == 'enseignant':
            flash('Conflit : L\'enseignant a déjà un cours à cette heure.', 'danger')
        elif type_conflit == 'salle':
            flash('Conflit : La salle est déjà occupée à cette heure.', 'danger')
        elif type_conflit == 'filiere':
            flash('Conflit : La filière a déjà un cours à cette heure.', 'danger')
        return redirect(url_for('admin.gestion_cours'))
    
    # Créer le créneau (assignation)
    creneau_id = EmploiDuTemps.creer(cours_id, enseignant_id, salle_id, jour, heure_debut, heure_fin)
    
    if creneau_id:
        flash('Enseignant assigné au cours avec succès.', 'success')
    else:
        flash('Erreur lors de l\'assignation de l\'enseignant.', 'danger')
    
    return redirect(url_for('admin.gestion_cours'))

# ==================== GESTION DES CRÉNEAUX EDT ====================

@admin_bp.route('/creneaux')
@role_required(['administration', 'directeur'])
def gestion_creneaux():
    """
    Page de gestion des créneaux d'emploi du temps
    """
    creneaux = EmploiDuTemps.obtenir_tous()
    cours = Cours.obtenir_tous()
    enseignants = Enseignant.obtenir_tous()
    salles = Salle.obtenir_toutes()

    # Ensure variables are not None to prevent template errors
    if creneaux is None:
        creneaux = []
    if cours is None:
        cours = []
    if enseignants is None:
        enseignants = []
    if salles is None:
        salles = []

    return render_template('gest_edt/scheduler.html',
                         creneaux=creneaux,
                         cours=cours,
                         enseignants=enseignants,
                         salles=salles)

@admin_bp.route('/creneaux/ajouter', methods=['POST'])
@role_required(['administration', 'directeur'])
def ajouter_creneau():
    """
    Ajouter un nouveau créneau d'emploi du temps
    """
    cours_id = request.form.get('cours_id', '').strip()
    enseignant_id = request.form.get('enseignant_id', '').strip()
    salle_id = request.form.get('salle_id', '').strip()
    jour = request.form.get('jour', '').strip()
    heure_debut = request.form.get('heure_debut', '').strip()
    heure_fin = request.form.get('heure_fin', '').strip()
    
    # Validation
    if not all([cours_id, enseignant_id, salle_id, jour, heure_debut, heure_fin]):
        flash('Tous les champs sont obligatoires.', 'danger')
        return redirect(url_for('admin.gestion_creneaux'))
    
    try:
        cours_id = int(cours_id)
        enseignant_id = int(enseignant_id)
        salle_id = int(salle_id)
    except ValueError:
        flash('Données invalides.', 'danger')
        return redirect(url_for('admin.gestion_creneaux'))
    
    # Vérifier les conflits (RG01)
    conflit = EmploiDuTemps.verifier_conflit(enseignant_id, salle_id, jour, heure_debut, heure_fin, cours_id)
    
    if conflit:
        type_conflit = conflit['type']
        if type_conflit == 'enseignant':
            flash('Conflit détecté : L\'enseignant a déjà un cours à cette heure.', 'danger')
        elif type_conflit == 'salle':
            flash('Conflit détecté : La salle est déjà occupée à cette heure.', 'danger')
        elif type_conflit == 'filiere':
            flash('Conflit détecté : La filière a déjà un cours à cette heure.', 'danger')
        return redirect(url_for('admin.gestion_creneaux'))
    
    # Créer le créneau
    creneau_id = EmploiDuTemps.creer(cours_id, enseignant_id, salle_id, jour, heure_debut, heure_fin)
    
    if creneau_id:
        flash('Créneau ajouté avec succès.', 'success')
    else:
        flash('Erreur lors de l\'ajout du créneau.', 'danger')
    
    return redirect(url_for('admin.gestion_creneaux'))

@admin_bp.route('/creneaux/modifier/<int:creneau_id>', methods=['POST'])
@role_required(['administration', 'directeur'])
def modifier_creneau(creneau_id):
    """
    Modifier un créneau existant
    """
    cours_id = request.form.get('cours_id', '').strip()
    enseignant_id = request.form.get('enseignant_id', '').strip()
    salle_id = request.form.get('salle_id', '').strip()
    jour = request.form.get('jour', '').strip()
    heure_debut = request.form.get('heure_debut', '').strip()
    heure_fin = request.form.get('heure_fin', '').strip()
    
    # Validation
    if not all([cours_id, enseignant_id, salle_id, jour, heure_debut, heure_fin]):
        flash('Tous les champs sont obligatoires.', 'danger')
        return redirect(url_for('admin.gestion_creneaux'))
    
    try:
        cours_id = int(cours_id)
        enseignant_id = int(enseignant_id)
        salle_id = int(salle_id)
    except ValueError:
        flash('Données invalides.', 'danger')
        return redirect(url_for('admin.gestion_creneaux'))
    
    # Vérifier les conflits (RG01) en excluant le créneau actuel
    conflit = EmploiDuTemps.verifier_conflit(enseignant_id, salle_id, jour, heure_debut, heure_fin, cours_id, creneau_id)
    
    if conflit:
        type_conflit = conflit['type']
        if type_conflit == 'enseignant':
            flash('Conflit détecté : L\'enseignant a déjà un cours à cette heure.', 'danger')
        elif type_conflit == 'salle':
            flash('Conflit détecté : La salle est déjà occupée à cette heure.', 'danger')
        elif type_conflit == 'filiere':
            flash('Conflit détecté : La filière a déjà un cours à cette heure.', 'danger')
        return redirect(url_for('admin.gestion_creneaux'))
    
    # Modifier le créneau
    resultat = EmploiDuTemps.modifier(creneau_id, cours_id, enseignant_id, salle_id, jour, heure_debut, heure_fin)
    
    if resultat:
        flash('Créneau modifié avec succès.', 'success')
    else:
        flash('Erreur lors de la modification du créneau.', 'danger')
    
    return redirect(url_for('admin.gestion_creneaux'))

@admin_bp.route('/creneaux/supprimer/<int:creneau_id>', methods=['POST'])
@role_required(['administration', 'directeur'])
def supprimer_creneau(creneau_id):
    """
    Supprimer un créneau
    """
    resultat = EmploiDuTemps.supprimer(creneau_id)

    if resultat:
        flash('Créneau supprimé avec succès.', 'success')
    else:
        flash('Erreur lors de la suppression du créneau.', 'danger')

    return redirect(url_for('admin.gestion_creneaux'))

# ==================== GESTION DES ÉTUDIANTS ====================

@admin_bp.route('/etudiants')
@role_required(['administration', 'directeur'])
def gestion_etudiants():
    """
    Page de gestion des étudiants
    """
    etudiants = Etudiant.obtenir_tous()
    filieres = Filiere.obtenir_toutes()
    return render_template('admin/gestion_etudiants.html', etudiants=etudiants, filieres=filieres)

@admin_bp.route('/etudiants/ajouter', methods=['POST'])
@role_required(['administration'])
def ajouter_etudiant():
    """
    Ajouter un nouvel étudiant (Redirige vers gestion_utilisateurs)
    """
    flash('Veuillez utiliser le module "Gestion des Utilisateurs" pour ajouter un étudiant.', 'info')
    return redirect(url_for('admin.gestion_utilisateurs'))

@admin_bp.route('/etudiants/modifier/<int:etudiant_id>', methods=['POST'])
@role_required(['administration'])
def modifier_etudiant(etudiant_id):
    """
    Modifier un étudiant existant (Redirige vers gestion_utilisateurs)
    """
    flash('Veuillez utiliser le module "Gestion des Utilisateurs" pour modifier un étudiant.', 'info')
    return redirect(url_for('admin.gestion_utilisateurs'))

@admin_bp.route('/etudiants/supprimer/<int:etudiant_id>', methods=['POST'])
@role_required(['administration', 'directeur'])
def supprimer_etudiant(etudiant_id):
    """
    Supprimer un étudiant
    """
    resultat = Utilisateur.supprimer(etudiant_id)

    if resultat:
        flash('Étudiant supprimé avec succès.', 'success')
    else:
        flash('Erreur lors de la suppression de l\'étudiant.', 'danger')

    return redirect(url_for('admin.gestion_etudiants'))


# ==================== GESTION DES UTILISATEURS ====================

@admin_bp.route('/utilisateurs')
@role_required(['administration', 'sous_admin', 'directeur', 'ADMIN', 'SUPER_ADMIN'])
def gestion_utilisateurs():
    """
    Page de gestion des utilisateurs
    """
    # Récupérer les filtres
    role_filtre = request.args.get('role', '')
    filiere_filtre = request.args.get('filiere', '')
    recherche = request.args.get('recherche', '')
    
    # Filtrer les utilisateurs
    if role_filtre or filiere_filtre or recherche:
        utilisateurs = Utilisateur.filtrer(
            role=role_filtre if role_filtre else None,
            filiere_id=int(filiere_filtre) if filiere_filtre else None,
            recherche=recherche if recherche else None
        )
    else:
        utilisateurs = Utilisateur.obtenir_tous()
    
    # Récupérer les filières pour le filtre
    filieres = Filiere.obtenir_toutes()
    
    return render_template('super_admin/users_manage.html', 
                         utilisateurs=utilisateurs,
                         filieres=filieres,
                         role_filtre=role_filtre,
                         filiere_filtre=filiere_filtre,
                         recherche=recherche)

@admin_bp.route('/utilisateurs/ajouter', methods=['POST'])
@role_required(['administration', 'sous_admin', 'directeur', 'ADMIN', 'SUPER_ADMIN'])
def ajouter_utilisateur():
    """
    Ajouter un nouvel utilisateur
    """
    nom = request.form.get('nom', '').strip()
    prenom = request.form.get('prenom', '').strip()
    role = request.form.get('role', '').strip()
    matricule_manuel = request.form.get('matricule', '').strip()
    filiere_id = request.form.get('filiere_id', '').strip()
    specialite = request.form.get('specialite', '').strip()
    
    # Validation
    if not all([nom, prenom, role]):
        flash('Le nom, prénom et rôle sont obligatoires.', 'danger')
        return redirect(url_for('admin.gestion_utilisateurs'))
    
    # Générer ou utiliser le matricule
    if matricule_manuel:
        matricule = matricule_manuel
    else:
        matricule = generer_matricule(role)
    
    # Vérifier si le matricule existe déjà
    if Utilisateur.obtenir_par_matricule(matricule):
        flash('Ce matricule est déjà utilisé.', 'danger')
        return redirect(url_for('admin.gestion_utilisateurs'))
    
    # Ajouter email et password
    email = request.form.get('email', '').strip()
    password = request.form.get('password', '').strip()

    # Validation
    if not all([nom, prenom, role]):
        flash('Le nom, prénom et rôle sont obligatoires.', 'danger')
        return redirect(url_for('admin.gestion_utilisateurs'))

    # Vérifier hiérarchie: ne peut créer que des rôles <= son niveau
    from app.utils import ROLES_HIERARCHY
    niveau_user = ROLES_HIERARCHY.get(session.get('role', ''), 0)
    niveau_new = ROLES_HIERARCHY.get(role, 0)
    if niveau_new > niveau_user:
        flash('Vous n\'avez pas la permission de créer un utilisateur avec ce rôle.', 'danger')
        return redirect(url_for('admin.gestion_utilisateurs'))

    # Générer ou utiliser le matricule
    if matricule_manuel:
        matricule = matricule_manuel
    else:
        matricule = generer_matricule(role)

    # Vérifier si le matricule existe déjà
    if Utilisateur.obtenir_par_matricule(matricule):
        flash('Ce matricule est déjà utilisé.', 'danger')
        return redirect(url_for('admin.gestion_utilisateurs'))

    # Hash password si fourni
    from werkzeug.security import generate_password_hash
    password_hash = generate_password_hash(password) if password else None

    # Créer l'utilisateur avec nouveaux champs
    created_by_id = session['utilisateur_id']
    utilisateur_id = Utilisateur.creer_complet(
        email, password_hash, role, nom, prenom, matricule,
        created_by_id, filiere_id=int(filiere_id) if filiere_id else None,
        specialite=specialite or None
    )

    if utilisateur_id:
        # Audit
        from app.models import AuditUsage
        AuditUsage.creer(created_by_id, 'CREATE_USER', meta={
            'new_user_id': utilisateur_id,
            'role': role,
            'matricule': matricule
        })

        flash(f'Utilisateur {prenom} {nom} (Matricule: {matricule}) ajouté avec succès.', 'success')
    else:
        flash('Erreur lors de l\'ajout de l\'utilisateur.', 'danger')
    
    return redirect(url_for('admin.gestion_utilisateurs'))

@admin_bp.route('/utilisateurs/modifier/<int:utilisateur_id>', methods=['POST'])
@role_required(['administration', 'sous_admin', 'directeur', 'ADMIN', 'SUPER_ADMIN'])
def modifier_utilisateur(utilisateur_id):
    """
    Modifier un utilisateur existant
    """
    nom = request.form.get('nom', '').strip()
    prenom = request.form.get('prenom', '').strip()
    matricule = request.form.get('matricule', '').strip()
    role = request.form.get('role', '').strip()
    filiere_id = request.form.get('filiere_id', '').strip()
    specialite = request.form.get('specialite', '').strip()
    
    # Validation
    if not all([nom, prenom, matricule, role]):
        flash('Tous les champs sont obligatoires.', 'danger')
        return redirect(url_for('admin.gestion_utilisateurs'))
    
    # Vérifier si le matricule existe déjà (sauf pour cet utilisateur)
    utilisateur_existant = Utilisateur.obtenir_par_matricule(matricule)
    if utilisateur_existant and utilisateur_existant['id'] != utilisateur_id:
        flash('Ce matricule est déjà utilisé par un autre utilisateur.', 'danger')
        return redirect(url_for('admin.gestion_utilisateurs'))
    
    # Modifier l'utilisateur
    resultat = Utilisateur.modifier(utilisateur_id, nom, prenom, matricule, role)
    
    if resultat:
        # Mettre à jour le profil spécifique
        if role == 'etudiant' and filiere_id:
            Etudiant.modifier(utilisateur_id, int(filiere_id))
        elif role == 'enseignant':
            Enseignant.modifier(utilisateur_id, specialite or '')
        
        flash('Utilisateur modifié avec succès.', 'success')
    else:
        flash('Erreur lors de la modification de l\'utilisateur.', 'danger')
    
    return redirect(url_for('admin.gestion_utilisateurs'))

@admin_bp.route('/utilisateurs/supprimer/<int:utilisateur_id>', methods=['POST'])
@role_required(['administration', 'sous_admin', 'directeur'])
def supprimer_utilisateur(utilisateur_id):
    """
    Supprimer un utilisateur
    """
    resultat = Utilisateur.supprimer(utilisateur_id)
    
    if resultat:
        flash('Utilisateur supprimé avec succès.', 'success')
    else:
        flash('Erreur lors de la suppression de l\'utilisateur.', 'danger')
    
    return redirect(url_for('admin.gestion_utilisateurs'))


# ==================== GESTION DES PRÉSENCES ====================

@admin_bp.route('/presences')
@role_required(['administration', 'sous_admin', 'directeur'])
def gestion_presences():
    """
    Page de gestion des présences
    """
    # Récupérer la semaine actuelle ou celle spécifiée
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    date_selectionnee = datetime.strptime(date_str, '%Y-%m-%d')
    
    # Calculer le début et la fin de la semaine
    debut_semaine = date_selectionnee - timedelta(days=date_selectionnee.weekday())
    fin_semaine = debut_semaine + timedelta(days=6)
    
    # Calculer les dates pour la navigation
    semaine_precedente = debut_semaine - timedelta(days=7)
    semaine_suivante = debut_semaine + timedelta(days=7)
    
    # Récupérer tous les créneaux
    creneaux = EmploiDuTemps.obtenir_tous()
    
    # Récupérer les présences de la semaine
    presences = Presence.obtenir_toutes_presences(
        debut_semaine.strftime('%Y-%m-%d'),
        fin_semaine.strftime('%Y-%m-%d')
    )
    
    # Organiser les présences par créneau et date
    presences_dict = {}
    for p in presences or []:
        key = f"{p['creneau_id']}_{p['date_cours']}"
        presences_dict[key] = p
    
    return render_template('gest_pres/mark_attendance.html',
                         creneaux=creneaux,
                         presences_dict=presences_dict,
                         debut_semaine=debut_semaine,
                         fin_semaine=fin_semaine,
                         date_selectionnee=date_selectionnee,
                         semaine_precedente=semaine_precedente,
                         semaine_suivante=semaine_suivante)

@admin_bp.route('/presences/marquer', methods=['POST'])
@role_required(['administration', 'sous_admin', 'directeur'])
def marquer_presence():
    """
    Marquer la présence/absence pour un créneau
    """
    creneau_id = request.form.get('creneau_id', '').strip()
    enseignant_id = request.form.get('enseignant_id', '').strip()
    statut = request.form.get('statut', '').strip()
    date_cours = request.form.get('date_cours', '').strip()
    remarques = request.form.get('remarques', '').strip()
    
    # Validation
    if not all([creneau_id, enseignant_id, statut, date_cours]):
        flash('Tous les champs sont obligatoires.', 'danger')
        return redirect(url_for('admin.gestion_presences'))
    
    try:
        creneau_id = int(creneau_id)
        enseignant_id = int(enseignant_id)
    except ValueError:
        flash('Données invalides.', 'danger')
        return redirect(url_for('admin.gestion_presences'))
    
    # Marquer la présence
    from flask import session
    marque_par = session.get('utilisateur_id')
    
    presence_id = Presence.marquer_presence(
        creneau_id, enseignant_id, statut, date_cours, remarques, marque_par
    )
    
    if presence_id:
        # Mettre à jour les heures de l'enseignant
        Enseignant.mettre_a_jour_heures(enseignant_id)
        flash('Présence marquée avec succès.', 'success')
    else:
        flash('Erreur lors du marquage de la présence.', 'danger')
    
    return redirect(url_for('admin.gestion_presences', date=date_cours))

@admin_bp.route('/presences/statistiques')
@role_required(['administration', 'sous_admin', 'directeur'])
def statistiques_presences():
    """
    Page des statistiques de présence des enseignants
    """
    enseignants = Enseignant.obtenir_tous()
    
    # Calculer les statistiques pour chaque enseignant
    stats_enseignants = []
    for ens in enseignants or []:
        stats = Presence.obtenir_statistiques_enseignant(ens['id'])
        
        # Calculer le taux de présence
        total = stats['total_seances'] if stats else 0
        presents = stats['presents'] if stats else 0
        taux = (presents / total * 100) if total > 0 else 0
        
        stats_enseignants.append({
            'enseignant': ens,
            'stats': stats,
            'taux_presence': taux,
            'heures_prevues': ens.get('total_heures_prevues', 0),
            'heures_effectuees': ens.get('total_heures_effectuees', 0)
        })
    
    return render_template('gest_pres/dashboard.html',
                         stats_enseignants=stats_enseignants)


# ==================== GESTION DES NOTES ====================

@admin_bp.route('/notes')
@role_required(['administration', 'sous_admin', 'directeur'])
def gestion_notes():
    """
    Page de gestion globale des notes
    """
    # Récupérer les filtres
    filiere_id = request.args.get('filiere_id', type=int)
    cours_id = request.args.get('cours_id', type=int)
    
    # Récupérer toutes les filières et cours pour les filtres
    filieres = Filiere.obtenir_toutes()
    cours_liste = Cours.obtenir_tous()
    
    notes = []
    statistiques = None
    filiere_selectionnee = None
    cours_selectionne = None
    
    if filiere_id:
        filiere_selectionnee = Filiere.obtenir_par_id(filiere_id)
        notes = Note.obtenir_par_filiere(filiere_id)
    
    if cours_id:
        cours_selectionne = Cours.obtenir_par_id(cours_id)
        notes = Note.obtenir_par_cours(cours_id)
        statistiques = Note.calculer_moyenne_cours(cours_id)
    
    return render_template('directeur/validation_notes.html',
                         filieres=filieres,
                         cours_liste=cours_liste,
                         filiere_selectionnee=filiere_selectionnee,
                         cours_selectionne=cours_selectionne,
                         notes=notes,
                         statistiques=statistiques)

@admin_bp.route('/notes/modifier/<int:note_id>', methods=['POST'])
@role_required(['administration', 'sous_admin', 'directeur', 'ADMIN', 'SUPER_ADMIN'])
def modifier_note_admin(note_id):
    """
    Modifier une note (admin uniquement)
    """
    note_val = request.form.get('note', '').strip()
    coefficient = request.form.get('coefficient', '').strip()
    commentaire = request.form.get('commentaire', '').strip()
    filiere_id = request.form.get('filiere_id', '')
    cours_id = request.form.get('cours_id', '')
    
    try:
        note_existante = Note.obtenir_par_id(note_id)
        if not note_existante:
            flash('Note introuvable.', 'danger')
            return redirect(url_for('admin.gestion_notes'))
        
        note_float = float(note_val)
        coef_float = float(coefficient)
        
        if note_float < 0 or note_float > 20:
            flash('La note doit être entre 0 et 20.', 'danger')
            return redirect(url_for('admin.gestion_notes', filiere_id=filiere_id, cours_id=cours_id))
        
        if coef_float <= 0:
            flash('Le coefficient doit être supérieur à 0.', 'danger')
            return redirect(url_for('admin.gestion_notes', filiere_id=filiere_id, cours_id=cours_id))
        
        Note.modifier(note_id, note_float, coef_float, commentaire)

        # Audit modification note
        from app.models import AuditUsage
        AuditUsage.creer(session['utilisateur_id'], 'MODIFY_NOTE', meta={
            'note_id': note_id,
            'new_note': note_float,
            'coefficient': coef_float
        })

        flash('Note modifiée avec succès!', 'success')
        
    except ValueError:
        flash('Valeurs invalides.', 'danger')
    except Exception as e:
        flash(f'Erreur: {str(e)}', 'danger')
    
    return redirect(url_for('admin.gestion_notes', filiere_id=filiere_id, cours_id=cours_id))

@admin_bp.route('/notes/supprimer/<int:note_id>', methods=['POST'])
@role_required(['administration', 'directeur'])
def supprimer_note_admin(note_id):
    """
    Supprimer une note (admin uniquement)
    """
    filiere_id = request.form.get('filiere_id', '')
    cours_id = request.form.get('cours_id', '')
    
    try:
        Note.supprimer(note_id)
        flash('Note supprimée avec succès!', 'success')
    except Exception as e:
        flash(f'Erreur: {str(e)}', 'danger')
    
    return redirect(url_for('admin.gestion_notes', filiere_id=filiere_id, cours_id=cours_id))


# ==================== GESTION DES EXAMENS ====================

@admin_bp.route('/examens')
@role_required(['administration', 'sous_admin', 'directeur'])
def gestion_examens():
    """
    Page de gestion des examens dans l'emploi du temps
    """
    # Récupérer tous les créneaux de type examen
    requete = """
        SELECT 
            edt.*,
            c.nom_cours, c.type_cours,
            u.nom as enseignant_nom, u.prenom as enseignant_prenom,
            s.nom_salle,
            f.nom_filiere, f.niveau
        FROM EmploiDuTemps edt
        JOIN Cours c ON edt.cours_id = c.id
        JOIN Enseignants e ON edt.enseignant_id = e.utilisateur_id
        JOIN Utilisateurs u ON e.utilisateur_id = u.id
        JOIN Salles s ON edt.salle_id = s.id
        JOIN Filieres f ON c.filiere_id = f.id
        WHERE edt.type_creneau IN ('examen', 'controle', 'tp_note')
        ORDER BY edt.jour, edt.heure_debut
    """
    from app.db import executer_requete
    examens = executer_requete(requete, obtenir_resultats=True)
    
    # Récupérer les données nécessaires pour le formulaire
    cours = Cours.obtenir_tous()
    enseignants = Enseignant.obtenir_tous()
    salles = Salle.obtenir_toutes()
    filieres = Filiere.obtenir_toutes()
    
    return render_template('gest_exam/data_correction.html',
                         examens=examens,
                         cours=cours,
                         enseignants=enseignants,
                         salles=salles,
                         filieres=filieres)

@admin_bp.route('/examens/ajouter', methods=['POST'])
@role_required(['administration', 'sous_admin', 'directeur'])
def ajouter_examen():
    """
    Ajouter un examen dans l'emploi du temps
    """
    cours_id = request.form.get('cours_id')
    enseignant_id = request.form.get('enseignant_id')
    salle_id = request.form.get('salle_id')
    jour = request.form.get('jour')
    heure_debut = request.form.get('heure_debut')
    heure_fin = request.form.get('heure_fin')
    type_creneau = request.form.get('type_creneau', 'examen')
    duree_examen = request.form.get('duree_examen', '')
    coefficient_examen = request.form.get('coefficient_examen', '')
    
    # Validation
    if not all([cours_id, enseignant_id, salle_id, jour, heure_debut, heure_fin]):
        flash('Tous les champs obligatoires doivent être remplis.', 'danger')
        return redirect(url_for('admin.gestion_examens'))
    
    # Vérifier les conflits
    conflit = EmploiDuTemps.verifier_conflit(
        int(enseignant_id), int(salle_id), jour, 
        heure_debut, heure_fin, int(cours_id)
    )
    
    if conflit:
        flash(f'Conflit détecté: {conflit["type"]}. Impossible d\'ajouter l\'examen.', 'danger')
        return redirect(url_for('admin.gestion_examens'))
    
    # Créer le créneau d'examen
    from app.db import executer_requete
    requete = """
        INSERT INTO EmploiDuTemps 
        (cours_id, enseignant_id, salle_id, jour, heure_debut, heure_fin, type_creneau, duree_examen, coefficient_examen)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    params = (cours_id, enseignant_id, salle_id, jour, heure_debut, heure_fin, 
              type_creneau, duree_examen if duree_examen else None, 
              coefficient_examen if coefficient_examen else None)
    
    resultat = executer_requete(requete, params)
    
    if resultat:
        flash('Examen ajouté avec succès dans l\'emploi du temps.', 'success')
    else:
        flash('Erreur lors de l\'ajout de l\'examen.', 'danger')
    
    return redirect(url_for('admin.gestion_examens'))


# ==================== GESTION DES CONFLITS ====================

@admin_bp.route('/conflits')
@role_required(['administration', 'sous_admin', 'directeur'])
def gestion_conflits():
    """
    Page de gestion des conflits de planification
    """
    conflits = Conflit.obtenir_conflits_actifs()
    
    return render_template('admin/conflict_check.html', conflits=conflits)

@admin_bp.route('/conflits/detecter', methods=['POST'])
@role_required(['administration', 'sous_admin', 'directeur'])
def detecter_conflits():
    """
    Détecter tous les conflits dans l'emploi du temps
    """
    conflits_detectes = Conflit.detecter_conflits()
    
    # Enregistrer les nouveaux conflits
    nb_nouveaux = 0
    for conflit_data in conflits_detectes:
        creneau_id = conflit_data['creneau_id']
        conflit = conflit_data['conflit']
        
        type_conflit = conflit['type']
        details = conflit['details']
        
        # Créer la description
        if type_conflit == 'enseignant':
            description = f"L'enseignant a déjà un cours le {details['jour']} de {details['heure_debut']} à {details['heure_fin']}"
        elif type_conflit == 'salle':
            description = f"La salle est déjà occupée le {details['jour']} de {details['heure_debut']} à {details['heure_fin']}"
        elif type_conflit == 'filiere':
            description = f"La filière a déjà un cours le {details['jour']} de {details['heure_debut']} à {details['heure_fin']}"
        else:
            description = "Conflit détecté"
        
        # Action suggérée
        action_suggeree = "Modifier l'horaire, changer la salle ou l'enseignant"
        
        # Enregistrer le conflit
        Conflit.enregistrer_conflit(
            creneau_id, type_conflit, details['id'],
            description, 'moyenne', action_suggeree
        )
        nb_nouveaux += 1
    
    flash(f'{nb_nouveaux} conflit(s) détecté(s) et enregistré(s).', 'info')
    return redirect(url_for('admin.gestion_conflits'))

@admin_bp.route('/conflits/resoudre/<int:conflit_id>', methods=['POST'])
@role_required(['administration', 'sous_admin', 'directeur'])
def resoudre_conflit(conflit_id):
    """
    Marquer un conflit comme résolu
    """
    utilisateur_id = session.get('utilisateur_id')
    Conflit.resoudre_conflit(conflit_id, utilisateur_id)
    
    flash('Conflit marqué comme résolu.', 'success')
    return redirect(url_for('admin.gestion_conflits'))

@admin_bp.route('/conflits/ignorer/<int:conflit_id>', methods=['POST'])
@role_required(['administration', 'sous_admin', 'directeur'])
def ignorer_conflit(conflit_id):
    """
    Marquer un conflit comme ignoré
    """
    Conflit.ignorer_conflit(conflit_id)
    
    flash('Conflit ignoré.', 'info')
    return redirect(url_for('admin.gestion_conflits'))


# ==================== GESTION DES PARENTS ====================

@admin_bp.route('/parents')
@role_required(['administration', 'sous_admin', 'directeur'])
def gestion_parents():
    """
    Page de gestion des parents et leurs liaisons avec les étudiants
    """
    # Récupérer tous les parents
    parents = Utilisateur.filtrer(role='parent')
    
    # Récupérer tous les étudiants
    etudiants = Etudiant.obtenir_tous()
    
    # Récupérer toutes les liaisons
    from app.db import executer_requete
    requete = "SELECT * FROM vue_parents_etudiants ORDER BY parent_nom, etudiant_nom"
    liaisons = executer_requete(requete, obtenir_resultats=True)
    
    return render_template('admin/gestion_parents.html',
                         parents=parents,
                         etudiants=etudiants,
                         liaisons=liaisons)

@admin_bp.route('/parents/lier', methods=['POST'])
@role_required(['administration', 'sous_admin', 'directeur'])
def lier_parent_etudiant():
    """
    Lier un parent à un étudiant
    """
    parent_id = request.form.get('parent_id')
    etudiant_id = request.form.get('etudiant_id')
    relation = request.form.get('relation', 'autre')
    telephone = request.form.get('telephone', '')
    email = request.form.get('email', '')
    adresse = request.form.get('adresse', '')
    contact_prioritaire = request.form.get('contact_prioritaire') == 'on'

    if not all([parent_id, etudiant_id]):
        flash('Le parent et l\'étudiant sont obligatoires.', 'danger')
        return redirect(url_for('admin.gestion_parents'))

    try:
        Parent.lier_etudiant(
            int(parent_id), int(etudiant_id), relation,
            telephone, email, adresse, contact_prioritaire
        )
        flash('Parent lié à l\'étudiant avec succès.', 'success')
    except Exception as e:
        flash(f'Erreur: {str(e)}', 'danger')

    return redirect(url_for('admin.gestion_parents'))

# ==================== IMPORT DE NOTES (WEB INTERFACE) ====================

@admin_bp.route('/import-notes')
@role_required(['administration', 'sous_admin', 'directeur', 'ADMIN', 'SUPER_ADMIN', 'ENSEIGNANT', 'GESTIONNAIRE_PV'])
def import_notes_page():
    """
    Page web pour l'import de notes
    """
    cours_liste = Cours.obtenir_tous()
    filieres = Filiere.obtenir_toutes()
    return render_template('gest_exam/import_upload.html', cours_liste=cours_liste, filieres=filieres)

@admin_bp.route('/import-notes/upload', methods=['POST'])
@role_required(['administration', 'sous_admin', 'directeur', 'ADMIN', 'SUPER_ADMIN', 'ENSEIGNANT', 'GESTIONNAIRE_PV'])
def importer_notes_web():
    """
    Importer des notes depuis l'interface web
    """
    if 'file' not in request.files:
        flash('Fichier requis', 'danger')
        return redirect(url_for('admin.import_notes_page'))

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('Format de fichier invalide (requis: .xlsx)', 'danger')
        return redirect(url_for('admin.import_notes_page'))

    cours_id = request.form.get('cours_id')
    filiere_id = request.form.get('filiere_id')

    if not all([cours_id, filiere_id]):
        flash('cours_id et filiere_id requis', 'danger')
        return redirect(url_for('admin.import_notes_page'))

    try:
        # Charger le fichier Excel
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        notes_importees = 0
        erreurs = []

        # Parcourir les lignes (ignorer l'en-tête)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Ligne vide
                continue

            try:
                # Format attendu: Matiere, Eleve, Note, Coefficient, TypeEvaluation, Date
                matiere_nom = str(row[0]).strip()
                eleve_matricule = str(row[1]).strip()
                note_val = float(row[2])
                coefficient = float(row[3]) if row[3] else 1.0
                type_eval = str(row[4]).strip() if row[4] else 'CC'
                date_eval = str(row[5]).strip() if len(row) > 5 and row[5] else None

                # Validation
                if not (0 <= note_val <= 20):
                    erreurs.append(f'Note invalide pour {eleve_matricule}: {note_val}')
                    continue
                if coefficient <= 0:
                    erreurs.append(f'Coefficient invalide pour {eleve_matricule}: {coefficient}')
                    continue

                # Résoudre cours_id depuis nom matière (simplifié)
                cours_id_resolved = int(cours_id)  # Pour l'instant, utiliser celui fourni

                # Résoudre etudiant_id depuis matricule
                etudiant = Utilisateur.obtenir_par_matricule(eleve_matricule)
                if not etudiant:
                    erreurs.append(f'Étudiant non trouvé: {eleve_matricule}')
                    continue
                etudiant_id = etudiant['id']

                # Créer la note
                Note.creer(
                    etudiant_id, cours_id_resolved, note_val, coefficient,
                    type_eval, date_eval, session['utilisateur_id']
                )
                notes_importees += 1

            except (ValueError, TypeError) as e:
                erreurs.append(f'Erreur ligne {ws.iter_rows().index(row) + 1}: {str(e)}')
                continue

        # Enregistrer l'import
        role_initiateur = session.get('role')
        import_id = ImportNote.creer(
            int(cours_id), int(filiere_id), session['utilisateur_id'],
            file.filename, notes_importees, role_initiateur
        )

        # Audit (commenté car AuditUsage n'existe pas)
        # AuditUsage.creer(session['utilisateur_id'], 'IMPORT_NOTES', meta={
        #     'import_id': import_id,
        #     'nb_lignes': notes_importees,
        #     'fichier': file.filename,
        #     'erreurs': len(erreurs)
        # })

        result = {
            'success': True,
            'imported': notes_importees,
            'errors': erreurs
        }

        flash(f'{notes_importees} notes importées avec succès', 'success')
        if erreurs:
            flash(f'Erreurs: {len(erreurs)}', 'warning')

    except Exception as e:
        result = {'success': False, 'errors': [f'Erreur traitement fichier: {str(e)}']}
        flash(f'Erreur traitement fichier: {str(e)}', 'danger')

    cours_liste = Cours.obtenir_tous()
    filieres = Filiere.obtenir_toutes()
    return render_template('gest_exam/import_upload.html', cours_liste=cours_liste, filieres=filieres, result=result)

# ==================== GÉNÉRATION DE BULLETINS (WEB INTERFACE) ====================

@admin_bp.route('/generer-bulletin')
@role_required(['GESTIONNAIRE_PV', 'administration', 'sous_admin', 'directeur', 'ADMIN', 'SUPER_ADMIN'])
def generer_bulletin_page():
    """
    Page web pour la génération de bulletins
    """
    etudiants = Etudiant.obtenir_tous()
    return render_template('gest_pv/generate_pdf.html', etudiants=etudiants)

@admin_bp.route('/generer-bulletin/generate', methods=['POST'])
@role_required(['GESTIONNAIRE_PV', 'administration', 'sous_admin', 'directeur', 'ADMIN', 'SUPER_ADMIN'])
def generer_bulletin_web():
    """
    Générer un bulletin PDF depuis l'interface web
    """
    etudiant_id = request.form.get('etudiant_id')
    periode = request.form.get('periode', 'S1')

    if not etudiant_id:
        flash('Étudiant requis', 'danger')
        return redirect(url_for('admin.generer_bulletin_page'))

    try:
        etudiant_id = int(etudiant_id)
    except ValueError:
        flash('ID étudiant invalide', 'danger')
        return redirect(url_for('admin.generer_bulletin_page'))

    # Récupérer les données de l'étudiant
    etudiant = Utilisateur.obtenir_par_id(etudiant_id)
    if not etudiant or etudiant['role'] not in ['ETUDIANT', 'etudiant']:
        flash('Étudiant non trouvé', 'danger')
        return redirect(url_for('admin.generer_bulletin_page'))

    # Récupérer la filière
    filiere = Filiere.obtenir_par_id(etudiant.get('filiere_id'))
    if not filiere:
        flash('Filière non trouvée', 'danger')
        return redirect(url_for('admin.generer_bulletin_page'))

    # Récupérer les notes
    notes = Note.obtenir_par_etudiant(etudiant_id)
    if not notes:
        flash('Aucune note trouvée pour cet étudiant', 'warning')
        etudiants = Etudiant.obtenir_tous()
        return render_template('admin/generer_bulletin.html', etudiants=etudiants)

    # Calculer moyennes par matière
    moyennes_matieres = {}
    for note in notes:
        matiere = note['nom_cours']
        if matiere not in moyennes_matieres:
            moyennes_matieres[matiere] = {'notes': [], 'coefficients': []}
        moyennes_matieres[matiere]['notes'].append(note['note'] * note['coefficient'])
        moyennes_matieres[matiere]['coefficients'].append(note['coefficient'])

    # Calculer moyennes
    data = [['Matière', 'Note', 'Coefficient', 'Moyenne']]
    moyenne_generale = 0
    total_coef = 0

    for matiere, details in moyennes_matieres.items():
        notes_ponderees = sum(details['notes'])
        coef_total = sum(details['coefficients'])
        moyenne = notes_ponderees / coef_total if coef_total > 0 else 0

        data.append([matiere, f"{notes_ponderees:.2f}", f"{coef_total}", f"{moyenne:.2f}"])
        moyenne_generale += moyenne * coef_total
        total_coef += coef_total

    moyenne_generale = moyenne_generale / total_coef if total_coef > 0 else 0

    # Générer PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # En-tête
    elements.append(Paragraph("UNIVERSITE INTERNATIONALE SCIENTIFIQUE ET TECHNIQUE", styles['Heading1']))
    elements.append(Paragraph("BULLETIN DE NOTES", styles['Heading2']))
    elements.append(Spacer(1, 12))

    # Infos étudiant
    elements.append(Paragraph(f"Étudiant: {etudiant['prenom']} {etudiant['nom']}", styles['Normal']))
    elements.append(Paragraph(f"Matricule: {etudiant['matricule']}", styles['Normal']))
    elements.append(Paragraph(f"Filière: {filiere['nom_filiere']}", styles['Normal']))
    elements.append(Paragraph(f"Période: {periode}", styles['Normal']))
    elements.append(Paragraph(f"Date: {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']))
    elements.append(Spacer(1, 12))

    # Tableau des notes
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    elements.append(Spacer(1, 12))

    # Moyenne générale
    elements.append(Paragraph(f"Moyenne Générale: {moyenne_generale:.2f}/20", styles['Heading3']))

    doc.build(elements)

    # Enregistrer la génération (commenté car BulletinGeneration n'existe pas)
    fichier_nom = f"bulletin_{etudiant['matricule']}_{periode}_{datetime.now().strftime('%Y%m%d')}.pdf"
    # BulletinGeneration.creer(etudiant_id, filiere['id'], periode, session['utilisateur_id'], fichier_nom)

    # Audit (commenté car AuditUsage n'existe pas)
    # AuditUsage.creer(session['utilisateur_id'], 'EXPORT_BULLETIN_PDF', meta={
    #     'etudiant_id': etudiant_id,
    #     'periode': periode,
    #     'fichier': fichier_nom
    # })

    buffer.seek(0)
    response = Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={fichier_nom}'}
    )

    flash('Bulletin généré avec succès', 'success')
    return response

# ==================== RAPPORT D'USAGE (WEB INTERFACE) ====================

@admin_bp.route('/usage-report')
@role_required(['SUPER_ADMIN'])
def usage_report_page():
    """
    Page web pour le rapport d'usage
    """
    return render_template('admin/usage_report.html')

@admin_bp.route('/usage-report/generate', methods=['POST'])
@role_required(['SUPER_ADMIN'])
def generer_rapport_usage_web():
    """
    Générer un rapport d'usage depuis l'interface web
    """
    format_type = request.form.get('format', 'json')
    debut = request.form.get('start')
    fin = request.form.get('end')

    # Statistiques utilisateurs
    stats_roles = Utilisateur.compter_par_role()

@admin_bp.route('/notes/valider/<int:note_id>', methods=['POST'])
@role_required(['directeur', 'administration', 'ADMIN', 'SUPER_ADMIN'])
def valider_note(note_id):
    """
    Valider une note pour la rendre visible aux étudiants et parents
    """
    from app.models import Note
    from flask import session
    user_id = session.get('user_id')
    if not user_id:
        flash('Session expirée. Veuillez vous reconnecter.', 'danger')
        return redirect(url_for('auth.connexion'))

    resultat = Note.valider_note(note_id, user_id)
    if resultat > 0:
        flash('Note validée avec succès. Elle est maintenant visible pour les étudiants et parents.', 'success')
    else:
        flash('Erreur lors de la validation de la note. Vérifiez que la note est en attente de validation.', 'danger')
    return redirect(url_for('directeur.tableau_de_bord'))

    # Statistiques bulletins (commenté car BulletinGeneration n'existe pas)
    # bulletins_stats = BulletinGeneration.stats()
    bulletins_stats = []

    # Statistiques imports
    imports_count = len(ImportNote.obtenir_historique())

    # Statistiques audit (commenté car AuditUsage n'existe pas)
    # audit_stats = AuditUsage.statistiques_actions(debut, fin)
    audit_stats = []

    data = {
        'generated_at': datetime.now().isoformat(),
        'period': {'start': debut, 'end': fin},
        'users_by_role': stats_roles,
        'active_users': len([u for u in utilisateurs_actifs if u['effective_last_login']]),
        'bulletins_generated': len(bulletins_stats) if bulletins_stats else 0,
        'notes_imports': imports_count,
        'audit_actions': audit_stats
    }

    if format_type == 'csv':
        import csv
        output = io.StringIO()
        writer = csv.writer(output)

        # En-têtes
        writer.writerow(['Métrique', 'Valeur'])

        # Données
        writer.writerow(['Date génération', data['generated_at']])
        writer.writerow(['Utilisateurs actifs', data['active_users']])
        writer.writerow(['Bulletins générés', data['bulletins_generated']])
        writer.writerow(['Imports de notes', data['notes_imports']])

        for role_stat in stats_roles:
            writer.writerow([f'Rôle {role_stat["role"]}', role_stat['count']])

        output.seek(0)
        response = Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment; filename=usage_report.csv'}
        )
        return response

    elif format_type == 'pdf':
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("RAPPORT D'USAGE DU SYSTÈME", styles['Heading1']))
        elements.append(Spacer(1, 12))

        # Statistiques générales
        elements.append(Paragraph("Statistiques Générales", styles['Heading2']))
        elements.append(Paragraph(f"Date de génération: {data['generated_at']}", styles['Normal']))
        elements.append(Paragraph(f"Utilisateurs actifs: {data['active_users']}", styles['Normal']))
        elements.append(Paragraph(f"Bulletins générés: {data['bulletins_generated']}", styles['Normal']))
        elements.append(Paragraph(f"Imports de notes: {data['notes_imports']}", styles['Normal']))
        elements.append(Spacer(1, 12))

        # Utilisateurs par rôle
        elements.append(Paragraph("Utilisateurs par Rôle", styles['Heading2']))
        role_data = [['Rôle', 'Nombre']]
        for stat in stats_roles:
            role_data.append([stat['role'], str(stat['count'])])
        role_table = Table(role_data)
        role_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(role_table)

        doc.build(elements)
        buffer.seek(0)
        response = Response(
            buffer.getvalue(),
            mimetype='application/pdf',
            headers={'Content-Disposition': 'attachment; filename=usage_report.pdf'}
        )
        return response

    # Audit (commenté car AuditUsage n'existe pas)
    # AuditUsage.creer(session['utilisateur_id'], 'USAGE_REPORT', meta={
    #     'format': format_type,
    #     'period': f"{debut} to {fin}" if debut and fin else 'all'
    # })

    flash('Rapport généré avec succès', 'success')
    return render_template('admin/usage_report.html', report_data=data)
