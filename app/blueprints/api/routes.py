"""
Routes API pour UIST-2ITS 
Endpoints REST pour gestion des utilisateurs, notes, bulletins, messages
Système de validation temps réel
"""
from flask import Blueprint, request, jsonify, session, Response
from app.models import Utilisateur, Note, ImportNote, Cours, Filiere, Message, Bulletin, AuditUsage
from app.utils import role_required, role_requis, generer_matricule
from werkzeug.security import generate_password_hash
import json
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
import openpyxl
from datetime import datetime

api_bp = Blueprint('api', __name__)

# ==================== GESTION DES UTILISATEURS ====================

@api_bp.route('/users', methods=['POST'])
@role_required(['administration', 'sous_admin', 'ADMIN', 'SUPER_ADMIN'])
def creer_utilisateur():
    """
    Créer un nouvel utilisateur via API
    """
    data = request.get_json()

    # Validation des champs requis
    required_fields = ['nom', 'prenom', 'role']
    for field in required_fields:
        if field not in data or not data[field].strip():
            return jsonify({'error': f'Champ {field} requis'}), 400

    nom = data['nom'].strip()
    prenom = data['prenom'].strip()
    role = data['role'].strip()
    email = data.get('email', '').strip()
    password = data.get('password', '').strip()
    matricule = data.get('matricule', '').strip()
    filiere_id = data.get('filiere_id')
    specialite = data.get('specialite', '').strip()

    # Vérifier hiérarchie des rôles
    from app.utils import ROLES_HIERARCHY
    niveau_user = ROLES_HIERARCHY.get(session.get('role', ''), 0)
    niveau_new = ROLES_HIERARCHY.get(role, 0)
    if niveau_new > niveau_user:
        return jsonify({'error': 'Permission insuffisante pour créer ce rôle'}), 403

    # Validation du rôle
    roles_valides = ['SUPER_ADMIN', 'ADMIN', 'GESTIONNAIRE_PV', 'ENSEIGNANT', 'ETUDIANT', 'PARENT']
    if role not in roles_valides:
        return jsonify({'error': 'Rôle invalide'}), 400

    # Validation filière pour étudiant
    if role == 'ETUDIANT' and not filiere_id:
        return jsonify({'error': 'filiere_id requis pour les étudiants'}), 400

    # Générer matricule si non fourni
    if not matricule:
        matricule = generer_matricule(role)

    # Vérifier unicité matricule
    if Utilisateur.obtenir_par_matricule(matricule):
        return jsonify({'error': 'Matricule déjà utilisé'}), 400

    # Hash password
    password_hash = generate_password_hash(password) if password else None

    # Créer utilisateur
    created_by_id = session['utilisateur_id']
    utilisateur_id = Utilisateur.creer_complet(
        email, password_hash, role, nom, prenom, matricule,
        created_by_id, filiere_id, specialite
    )

    if not utilisateur_id:
        return jsonify({'error': 'Erreur lors de la création'}), 500

    # Audit (commenté car AuditUsage n'existe pas)
    # AuditUsage.creer(created_by_id, 'CREATE_USER', meta={
    #     'new_user_id': utilisateur_id,
    #     'role': role,
    #     'matricule': matricule,
    #     'via_api': True
    # })

    return jsonify({
        'id': utilisateur_id,
        'role': role,
        'matricule': matricule,
        'message': 'Utilisateur créé avec succès'
    }), 201

# ==================== IMPORT DE NOTES ====================

@api_bp.route('/notes/import', methods=['POST'])
@role_requis('administration', 'sous_admin', 'ADMIN', 'SUPER_ADMIN', 'ENSEIGNANT')
def importer_notes():
    """
    Importer des notes depuis un fichier Excel
    """
    if 'file' not in request.files:
        return jsonify({'error': 'Fichier requis'}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Format de fichier invalide (requis: .xlsx)'}), 400

    cours_id = request.form.get('cours_id')
    filiere_id = request.form.get('filiere_id')

    if not all([cours_id, filiere_id]):
        return jsonify({'error': 'cours_id et filiere_id requis'}), 400

    try:
        # Charger le fichier Excel
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        notes_importees = 0
        erreurs = []

        # Parcourir les lignes (ignorer l'en-tête)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Ligne vide
                continue

            try:
                # Format attendu: Matiere, Eleve, Note, Coefficient, TypeEvaluation, Date
                matiere_nom = str(row[0]).strip()
                eleve_matricule = str(row[1]).strip()
                note_val = float(row[2])
                coefficient = float(row[3]) if row[3] else 1.0
                type_eval = str(row[4]).strip() if row[4] else 'CC'
                date_eval = str(row[5]).strip() if len(row) > 5 and row[5] else None

                # Validation
                if not (0 <= note_val <= 20):
                    erreurs.append(f'Note invalide pour {eleve_matricule}: {note_val}')
                    continue
                if coefficient <= 0:
                    erreurs.append(f'Coefficient invalide pour {eleve_matricule}: {coefficient}')
                    continue

                # Résoudre cours_id depuis nom matière (simplifié)
                # TODO: Implémenter résolution par nom matière
                cours_id_resolved = int(cours_id)  # Pour l'instant, utiliser celui fourni

                # Résoudre etudiant_id depuis matricule
                etudiant = Utilisateur.obtenir_par_matricule(eleve_matricule)
                if not etudiant:
                    erreurs.append(f'Étudiant non trouvé: {eleve_matricule}')
                    continue
                etudiant_id = etudiant['id']

                # Créer la note
                Note.creer(
                    etudiant_id, cours_id_resolved, note_val, coefficient,
                    type_eval, date_eval, session['utilisateur_id']
                )
                notes_importees += 1

            except (ValueError, TypeError) as e:
                erreurs.append(f'Erreur ligne {ws.iter_rows().index(row) + 1}: {str(e)}')
                continue

        # Enregistrer l'import
        role_initiateur = session.get('role')
        import_id = ImportNote.creer(
            int(cours_id), int(filiere_id), session['utilisateur_id'],
            file.filename, notes_importees, role_initiateur
        )

        # Audit (commenté car AuditUsage n'existe pas)
        # AuditUsage.creer(session['utilisateur_id'], 'IMPORT_NOTES', meta={
        #     'import_id': import_id,
        #     'nb_lignes': notes_importees,
        #     'fichier': file.filename,
        #     'erreurs': len(erreurs)
        # })

        return jsonify({
            'success': True,
            'imported': notes_importees,
            'errors': erreurs,
            'message': f'{notes_importees} notes importées'
        }), 200

    except Exception as e:
        return jsonify({'error': f'Erreur traitement fichier: {str(e)}'}), 500

# ==================== GÉNÉRATION DE BULLETINS ====================

@api_bp.route('/bulletins/<int:etudiant_id>', methods=['GET'])
@role_requis('GESTIONNAIRE_PV', 'administration', 'sous_admin', 'ADMIN', 'SUPER_ADMIN')
def telecharger_bulletin_pdf(etudiant_id):
    """
    Générer un bulletin PDF pour un étudiant
    """
    periode = request.args.get('periode', 'S1')

    # Récupérer les données de l'étudiant
    etudiant = Utilisateur.obtenir_par_id(etudiant_id)
    if not etudiant or etudiant['role'] not in ['ETUDIANT', 'etudiant']:
        return jsonify({'error': 'Étudiant non trouvé'}), 404

    # Récupérer la filière
    filiere = Filiere.obtenir_par_id(etudiant.get('filiere_id'))
    if not filiere:
        return jsonify({'error': 'Filière non trouvée'}), 404

    # Récupérer les notes
    notes = Note.obtenir_par_etudiant(etudiant_id)
    if not notes:
        return jsonify({'error': 'Aucune note trouvée'}), 404

    # Calculer moyennes par matière
    moyennes_matieres = {}
    for note in notes:
        matiere = note['nom_cours']
        if matiere not in moyennes_matieres:
            moyennes_matieres[matiere] = {'notes': [], 'coefficients': []}
        moyennes_matieres[matiere]['notes'].append(note['note'] * note['coefficient'])
        moyennes_matieres[matiere]['coefficients'].append(note['coefficient'])

    # Calculer moyennes
    data = [['Matière', 'Note', 'Coefficient', 'Moyenne']]
    moyenne_generale = 0
    total_coef = 0

    for matiere, details in moyennes_matieres.items():
        notes_ponderees = sum(details['notes'])
        coef_total = sum(details['coefficients'])
        moyenne = notes_ponderees / coef_total if coef_total > 0 else 0

        data.append([matiere, f"{notes_ponderees:.2f}", f"{coef_total}", f"{moyenne:.2f}"])
        moyenne_generale += moyenne * coef_total
        total_coef += coef_total

    moyenne_generale = moyenne_generale / total_coef if total_coef > 0 else 0

    # Générer PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # En-tête
    elements.append(Paragraph("UNIVERSITE INTERNATIONALE SCIENTIFIQUE ET TECHNIQUE", styles['Heading1']))
    elements.append(Paragraph("BULLETIN DE NOTES", styles['Heading2']))
    elements.append(Spacer(1, 12))

    # Infos étudiant
    elements.append(Paragraph(f"Étudiant: {etudiant['prenom']} {etudiant['nom']}", styles['Normal']))
    elements.append(Paragraph(f"Matricule: {etudiant['matricule']}", styles['Normal']))
    elements.append(Paragraph(f"Filière: {filiere['nom_filiere']}", styles['Normal']))
    elements.append(Paragraph(f"Période: {periode}", styles['Normal']))
    elements.append(Paragraph(f"Date: {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']))
    elements.append(Spacer(1, 12))

    # Tableau des notes
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    elements.append(Spacer(1, 12))

    # Moyenne générale
    elements.append(Paragraph(f"Moyenne Générale: {moyenne_generale:.2f}/20", styles['Heading3']))

    doc.build(elements)

    # Enregistrer la génération (commenté car BulletinGeneration n'existe pas)
    fichier_nom = f"bulletin_{etudiant['matricule']}_{periode}_{datetime.now().strftime('%Y%m%d')}.pdf"
    # BulletinGeneration.creer(etudiant_id, filiere['id'], periode, session['utilisateur_id'], fichier_nom)

    # Audit (commenté car AuditUsage n'existe pas)
    # AuditUsage.creer(session['utilisateur_id'], 'EXPORT_BULLETIN_PDF', meta={
    #     'etudiant_id': etudiant_id,
    #     'periode': periode,
    #     'fichier': fichier_nom
    # })

    buffer.seek(0)
    return Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={fichier_nom}'}
    )

# ==================== RAPPORT D'USAGE ====================

@api_bp.route('/admin/usage-report', methods=['GET'])
@role_required(['SUPER_ADMIN'])
def rapport_usage():
    """
    Générer un rapport d'usage du système
    """
    format_type = request.args.get('format', 'json')
    debut = request.args.get('start')
    fin = request.args.get('end')

    # Statistiques utilisateurs
    stats_roles = Utilisateur.compter_par_role()
    utilisateurs_actifs = Utilisateur.lister_avec_last_login()

    # Statistiques bulletins (commenté car BulletinGeneration n'existe pas)
    # bulletins_stats = BulletinGeneration.stats()
    bulletins_stats = []

    # Statistiques imports
    imports_count = len(ImportNote.obtenir_historique())

    # Statistiques audit (commenté car AuditUsage n'existe pas)
    # audit_stats = AuditUsage.statistiques_actions(debut, fin)
    audit_stats = []

    data = {
        'generated_at': datetime.now().isoformat(),
        'period': {'start': debut, 'end': fin},
        'users_by_role': stats_roles,
        'active_users': len([u for u in utilisateurs_actifs if u['effective_last_login']]),
        'bulletins_generated': len(bulletins_stats) if bulletins_stats else 0,
        'notes_imports': imports_count,
        'audit_actions': audit_stats
    }

    if format_type == 'csv':
        import csv
        output = io.StringIO()
        writer = csv.writer(output)

        # En-têtes
        writer.writerow(['Métrique', 'Valeur'])

        # Données
        writer.writerow(['Date génération', data['generated_at']])
        writer.writerow(['Utilisateurs actifs', data['active_users']])
        writer.writerow(['Bulletins générés', data['bulletins_generated']])
        writer.writerow(['Imports de notes', data['notes_imports']])

        for role_stat in stats_roles:
            writer.writerow([f'Rôle {role_stat["role"]}', role_stat['count']])

        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment; filename=usage_report.csv'}
        )

    elif format_type == 'pdf':
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("RAPPORT D'USAGE DU SYSTÈME", styles['Heading1']))
        elements.append(Spacer(1, 12))

        # Statistiques générales
        elements.append(Paragraph("Statistiques Générales", styles['Heading2']))
        elements.append(Paragraph(f"Date de génération: {data['generated_at']}", styles['Normal']))
        elements.append(Paragraph(f"Utilisateurs actifs: {data['active_users']}", styles['Normal']))
        elements.append(Paragraph(f"Bulletins générés: {data['bulletins_generated']}", styles['Normal']))
        elements.append(Paragraph(f"Imports de notes: {data['notes_imports']}", styles['Normal']))
        elements.append(Spacer(1, 12))

        # Utilisateurs par rôle
        elements.append(Paragraph("Utilisateurs par Rôle", styles['Heading2']))
        role_data = [['Rôle', 'Nombre']]
        for stat in stats_roles:
            role_data.append([stat['role'], str(stat['count'])])
        role_table = Table(role_data)
        role_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(role_table)

        doc.build(elements)
        buffer.seek(0)
        return Response(
            buffer.getvalue(),
            mimetype='application/pdf',
            headers={'Content-Disposition': 'attachment; filename=usage_report.pdf'}
        )

    # Audit
    AuditUsage.creer(session['utilisateur_id'], 'USAGE_REPORT', meta={
        'format': format_type,
        'period': f"{debut} to {fin}" if debut and fin else 'all'
    })


# ==================== API NOTES - WORKFLOW UNICAMPUS ====================

@api_bp.route('/notes/en-attente', methods=['GET'])
@role_required(['DIRECTEUR', 'ADMIN', 'SUPER_ADMIN'])
def obtenir_notes_en_attente():
    """
    Récupère toutes les notes en attente de validation (pour le Directeur)
    Polling endpoint - appelé toutes les 5 secondes
    """
    filiere_id = request.args.get('filiere_id', type=int)
    
    try:
        notes = Note.obtenir_notes_en_attente(filiere_id)
        
        # Formater les données pour le frontend
        notes_formatted = []
        for note in notes:
            notes_formatted.append({
                'id': note['id'],
                'etudiant': {
                    'id': note['etudiant_id'],
                    'nom': note['etudiant_nom'],
                    'prenom': note['etudiant_prenom'],
                    'matricule': note['matricule']
                },
                'cours': {
                    'id': note['cours_id'],
                    'nom': note['nom_cours'],
                    'type': note['type_cours']
                },
                'filiere': {
                    'id': note.get('filiere_id'),
                    'nom': note['nom_filiere'],
                    'niveau': note['niveau']
                },
                'type_evaluation': note['type_evaluation'],
                'note': float(note['note']),
                'coefficient': float(note['coefficient']),
                'date_evaluation': note['date_evaluation'].strftime('%Y-%m-%d') if note.get('date_evaluation') else None,
                'commentaire': note.get('commentaire', ''),
                'saisi_par': {
                    'nom': note['saisi_par_nom'],
                    'prenom': note['saisi_par_prenom']
                },
                'date_creation': note['date_creation'].strftime('%Y-%m-%d %H:%M:%S') if note.get('date_creation') else None
            })
        
        return jsonify({
            'success': True,
            'count': len(notes_formatted),
            'notes': notes_formatted
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@api_bp.route('/notes/valider/<int:note_id>', methods=['POST'])
@role_requis('DIRECTEUR', 'ADMIN', 'SUPER_ADMIN')
def valider_note(note_id):
    """
    Valide une note (change le statut à VALIDÉ)
    """
    try:
        directeur_id = session['utilisateur_id']
        
        # Vérifier que la note existe et est en attente
        note = Note.obtenir_par_id(note_id)
        if not note:
            return jsonify({
                'success': False,
                'error': 'Note non trouvée'
            }), 404
        
        if note['statut'] != 'EN_ATTENTE_DIRECTEUR':
            return jsonify({
                'success': False,
                'error': f'Note déjà {note["statut"]}'
            }), 400
        
        # Valider la note
        result = Note.valider_note(note_id, directeur_id)
        
        if result:
            # Audit
            AuditUsage.creer(directeur_id, 'VALIDATE_NOTE', meta={
                'note_id': note_id,
                'etudiant_id': note['etudiant_id'],
                'cours_id': note['cours_id']
            })
            
            return jsonify({
                'success': True,
                'message': 'Note validée avec succès'
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'Erreur lors de la validation'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@api_bp.route('/notes/modifier/<int:note_id>', methods=['PUT'])
@role_requis('DIRECTEUR', 'ADMIN', 'SUPER_ADMIN')
def modifier_note_non_validee(note_id):
    """
    Modifie une note non validée (réservé au Directeur)
    """
    try:
        data = request.get_json()
        
        # Vérifier que la note existe
        note = Note.obtenir_par_id(note_id)
        if not note:
            return jsonify({
                'success': False,
                'error': 'Note non trouvée'
            }), 404
        
        if note['statut'] == 'VALIDÉ':
            return jsonify({
                'success': False,
                'error': 'Impossible de modifier une note validée'
            }), 400
        
        # Récupérer les nouvelles valeurs
        nouvelle_note = data.get('note')
        nouveau_coefficient = data.get('coefficient')
        nouveau_commentaire = data.get('commentaire')
        
        if nouvelle_note is None:
            return jsonify({
                'success': False,
                'error': 'Nouvelle note requise'
            }), 400
        
        # Valider la note (0-20)
        try:
            nouvelle_note = float(nouvelle_note)
            if nouvelle_note < 0 or nouvelle_note > 20:
                raise ValueError
        except ValueError:
            return jsonify({
                'success': False,
                'error': 'Note doit être entre 0 et 20'
            }), 400
        
        # Modifier la note
        result = Note.modifier_note_non_validee(
            note_id, 
            nouvelle_note, 
            nouveau_coefficient, 
            nouveau_commentaire
        )
        
        if result:
            # Audit
            AuditUsage.creer(session['utilisateur_id'], 'MODIFY_NOTE', meta={
                'note_id': note_id,
                'old_value': float(note['note']),
                'new_value': nouvelle_note
            })
            
            return jsonify({
                'success': True,
                'message': 'Note modifiée avec succès'
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'Erreur lors de la modification'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@api_bp.route('/notes/etudiant/<int:etudiant_id>', methods=['GET'])
@role_required(['ETUDIANT', 'PARENT', 'ENSEIGNANT', 'ADMIN', 'SUPER_ADMIN', 'DIRECTEUR'])
def obtenir_notes_etudiant(etudiant_id):
    """
    Récupère les notes validées d'un étudiant
    Les étudiants ne voient que leurs propres notes
    Les parents voient les notes de leurs enfants
    """
    try:
        user_id = session['utilisateur_id']
        user_role = session['role']
        
        # Vérification des permissions
        if user_role == 'ETUDIANT':
            # Un étudiant ne peut voir que ses propres notes
            if user_id != etudiant_id:
                return jsonify({
                    'success': False,
                    'error': 'Accès non autorisé'
                }), 403
        
        elif user_role == 'PARENT':
            # Un parent ne peut voir que les notes de ses enfants
            from app.models import Parent
            enfants = Parent.obtenir_enfants(user_id)
            enfants_ids = [e['etudiant_id'] for e in enfants]
            
            if etudiant_id not in enfants_ids:
                return jsonify({
                    'success': False,
                    'error': 'Accès non autorisé'
                }), 403
        
        # Récupérer les notes validées
        notes = Note.obtenir_notes_validees_etudiant(etudiant_id)
        
        # Formater les données
        notes_formatted = []
        for note in notes:
            notes_formatted.append({
                'id': note['id'],
                'cours': {
                    'nom': note['nom_cours'],
                    'type': note['type_cours']
                },
                'filiere': {
                    'nom': note['nom_filiere'],
                    'niveau': note['niveau']
                },
                'type_evaluation': note['type_evaluation'],
                'note': float(note['note']),
                'coefficient': float(note['coefficient']),
                'date_evaluation': note['date_evaluation'].strftime('%Y-%m-%d') if note.get('date_evaluation') else None,
                'commentaire': note.get('commentaire', ''),
                'date_validation': note['date_validation'].strftime('%Y-%m-%d') if note.get('date_validation') else None
            })
        
        # Calculer la moyenne
        moyenne_data = Note.calculer_moyenne_etudiant(etudiant_id, seulement_validees=True)
        
        return jsonify({
            'success': True,
            'count': len(notes_formatted),
            'notes': notes_formatted,
            'moyenne': {
                'generale': float(moyenne_data['moyenne']) if moyenne_data and moyenne_data['moyenne'] else 0,
                'nb_notes': moyenne_data['nb_notes'] if moyenne_data else 0,
                'note_min': float(moyenne_data['note_min']) if moyenne_data and moyenne_data['note_min'] else 0,
                'note_max': float(moyenne_data['note_max']) if moyenne_data and moyenne_data['note_max'] else 0
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==================== API MESSAGES ====================

@api_bp.route('/messages/non-lus', methods=['GET'])
@role_required(['ETUDIANT', 'PARENT', 'ENSEIGNANT', 'ADMIN', 'SUPER_ADMIN', 'DIRECTEUR', 'GESTIONNAIRE_PV', 'GESTIONNAIRE_EXAMENS', 'GESTIONNAIRE_EDT'])
def obtenir_messages_non_lus():
    """
    Récupère les messages non lus de l'utilisateur connecté
    Polling endpoint
    """
    try:
        user_id = session['utilisateur_id']
        messages = Message.obtenir_messages_non_lus(user_id)
        
        # Formater les données
        messages_formatted = []
        for msg in messages:
            messages_formatted.append({
                'id': msg['id'],
                'expediteur': {
                    'nom': msg['expediteur_nom'],
                    'prenom': msg['expediteur_prenom'],
                    'role': msg['expediteur_role']
                },
                'type': msg['type_message'],
                'sujet': msg['sujet'],
                'contenu': msg['contenu'],
                'note_id': msg.get('note_id'),
                'date_creation': msg['date_creation'].strftime('%Y-%m-%d %H:%M:%S') if msg.get('date_creation') else None
            })
        
        return jsonify({
            'success': True,
            'count': len(messages_formatted),
            'messages': messages_formatted
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@api_bp.route('/messages/envoyer', methods=['POST'])
@role_required(['ETUDIANT', 'PARENT', 'ENSEIGNANT', 'ADMIN', 'SUPER_ADMIN', 'DIRECTEUR', 'GESTIONNAIRE_PV', 'GESTIONNAIRE_EXAMENS', 'GESTIONNAIRE_EDT'])
def envoyer_message():
    """
    Envoie un message
    """
    try:
        data = request.get_json()
        
        # Validation
        required_fields = ['destinataire_id', 'sujet', 'contenu']
        for field in required_fields:
            if field not in data:
                return jsonify({
                    'success': False,
                    'error': f'Champ {field} requis'
                }), 400
        
        expediteur_id = session['utilisateur_id']
        destinataire_id = data['destinataire_id']
        sujet = data['sujet'].strip()
        contenu = data['contenu'].strip()
        type_message = data.get('type_message', 'MESSAGE')
        note_id = data.get('note_id')
        
        # Créer le message
        message_id = Message.creer(
            expediteur_id,
            destinataire_id,
            sujet,
            contenu,
            type_message,
            note_id
        )
        
        if message_id:
            # Audit
            AuditUsage.creer(expediteur_id, 'SEND_MESSAGE', meta={
                'message_id': message_id,
                'destinataire_id': destinataire_id,
                'type': type_message
            })
            
            return jsonify({
                'success': True,
                'message_id': message_id,
                'message': 'Message envoyé avec succès'
            }), 201
        else:
            return jsonify({
                'success': False,
                'error': 'Erreur lors de l\'envoi'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@api_bp.route('/messages/signalement', methods=['POST'])
@role_required(['ETUDIANT'])
def creer_signalement():
    """
    Crée un signalement de note (envoyé au Directeur)
    """
    try:
        data = request.get_json()
        
        # Validation
        if 'note_id' not in data or 'contenu' not in data:
            return jsonify({
                'success': False,
                'error': 'note_id et contenu requis'
            }), 400
        
        etudiant_id = session['utilisateur_id']
        note_id = data['note_id']
        contenu = data['contenu'].strip()
        
        # Vérifier que la note appartient à l'étudiant
        note = Note.obtenir_par_id(note_id)
        if not note or note['etudiant_id'] != etudiant_id:
            return jsonify({
                'success': False,
                'error': 'Note non trouvée ou accès non autorisé'
            }), 404
        
        # Créer le signalement
        message_id = Message.creer_signalement(etudiant_id, note_id, contenu)
        
        if message_id:
            # Audit
            AuditUsage.creer(etudiant_id, 'CREATE_SIGNALEMENT', meta={
                'message_id': message_id,
                'note_id': note_id
            })
            
            return jsonify({
                'success': True,
                'message_id': message_id,
                'message': 'Signalement envoyé au Directeur'
            }), 201
        else:
            return jsonify({
                'success': False,
                'error': 'Erreur lors de la création du signalement'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@api_bp.route('/messages/marquer-lu/<int:message_id>', methods=['PUT'])
@role_requis('ETUDIANT', 'PARENT', 'ENSEIGNANT', 'ADMIN', 'SUPER_ADMIN', 'DIRECTEUR', 'GESTIONNAIRE_PV', 'GESTIONNAIRE_EXAMENS', 'GESTIONNAIRE_EDT')
def marquer_message_lu(message_id):
    """
    Marque un message comme lu
    """
    try:
        result = Message.marquer_comme_lu(message_id)
        
        if result:
            return jsonify({
                'success': True,
                'message': 'Message marqué comme lu'
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'Message non trouvé'
            }), 404
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==================== API BULLETINS ====================

@api_bp.route('/bulletins/generer', methods=['POST'])
@role_requis('GESTIONNAIRE_PV', 'ADMIN', 'SUPER_ADMIN')
def creer_bulletin():
    """
    Génère un bulletin PDF pour un étudiant
    """
    try:
        data = request.get_json()
        
        # Validation
        required_fields = ['etudiant_id', 'semestre', 'annee_academique']
        for field in required_fields:
            if field not in data:
                return jsonify({
                    'success': False,
                    'error': f'Champ {field} requis'
                }), 400
        
        etudiant_id = data['etudiant_id']
        semestre = data['semestre']
        annee_academique = data['annee_academique']
        genere_par = session['utilisateur_id']
        
        # Générer le bulletin
        bulletin_id = Bulletin.generer_bulletin(
            etudiant_id,
            semestre,
            annee_academique,
            genere_par
        )
        
        if bulletin_id:
            # Audit
            AuditUsage.creer(genere_par, 'GENERATE_BULLETIN', meta={
                'bulletin_id': bulletin_id,
                'etudiant_id': etudiant_id,
                'semestre': semestre
            })
            
            return jsonify({
                'success': True,
                'bulletin_id': bulletin_id,
                'message': 'Bulletin généré avec succès'
            }), 201
        else:
            return jsonify({
                'success': False,
                'error': 'Erreur lors de la génération'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@api_bp.route('/bulletins/etudiant/<int:etudiant_id>', methods=['GET'])
@role_requis('ETUDIANT', 'PARENT', 'GESTIONNAIRE_PV', 'ADMIN', 'SUPER_ADMIN')
def obtenir_bulletins_etudiant(etudiant_id):
    """
    Récupère tous les bulletins d'un étudiant
    """
    try:
        user_id = session['utilisateur_id']
        user_role = session['role']
        
        # Vérification des permissions
        if user_role == 'ETUDIANT' and user_id != etudiant_id:
            return jsonify({
                'success': False,
                'error': 'Accès non autorisé'
            }), 403
        
        elif user_role == 'PARENT':
            from app.models import Parent
            enfants = Parent.obtenir_enfants(user_id)
            enfants_ids = [e['etudiant_id'] for e in enfants]
            
            if etudiant_id not in enfants_ids:
                return jsonify({
                    'success': False,
                    'error': 'Accès non autorisé'
                }), 403
        
        # Récupérer les bulletins
        bulletins = Bulletin.obtenir_bulletins_etudiant(etudiant_id)
        
        # Formater les données
        bulletins_formatted = []
        for bulletin in bulletins:
            bulletins_formatted.append({
                'id': bulletin['id'],
                'semestre': bulletin['semestre'],
                'annee_academique': bulletin['annee_academique'],
                'moyenne_generale': float(bulletin['moyenne_generale']) if bulletin.get('moyenne_generale') else 0,
                'rang': bulletin.get('rang'),
                'appreciation': bulletin.get('appreciation', ''),
                'fichier_pdf': bulletin.get('fichier_pdf'),
                'date_generation': bulletin['date_generation'].strftime('%Y-%m-%d %H:%M:%S') if bulletin.get('date_generation') else None
            })
        
        return jsonify({
            'success': True,
            'count': len(bulletins_formatted),
            'bulletins': bulletins_formatted
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

    return jsonify(data)

# ==================== API PRESENCES ====================

@api_bp.route('/presences/enseignants', methods=['GET'])
@role_requis('administration', 'sous_admin', 'ADMIN', 'SUPER_ADMIN')
def obtenir_presences_enseignants():
    """
    Récupère les présences des enseignants avec filtres
    """
    try:
        date = request.args.get('date')
        enseignant_id = request.args.get('enseignant_id')
        statut = request.args.get('statut')

        # Récupérer les présences
        presences = Presence.obtenir_presences_enseignants(date, enseignant_id, statut)

        # Calculer les statistiques
        stats = Presence.calculer_stats_presences(date)

        return jsonify({
            'success': True,
            'presences': presences,
            'stats': stats
        }), 200

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@api_bp.route('/presences/form', methods=['GET'])
@role_requis('administration', 'sous_admin', 'ADMIN', 'SUPER_ADMIN')
def obtenir_formulaire_presence():
    """
    Récupère le formulaire pour marquer une présence
    """
    try:
        creneau_id = request.args.get('creneau_id')
        enseignant_id = request.args.get('enseignant_id')
        date = request.args.get('date')
        modification = request.args.get('modification', 'false').lower() == 'true'

        if not all([creneau_id, enseignant_id, date]):
            return jsonify({'error': 'Paramètres manquants'}), 400

        # Récupérer les données du créneau
        from app.models import EmploiDuTemps
        creneau = EmploiDuTemps.obtenir_par_id(int(creneau_id))
        if not creneau:
            return jsonify({'error': 'Créneau non trouvé'}), 404

        # Récupérer la présence existante si modification
        presence_existante = None
        if modification:
            presence_existante = Presence.obtenir_presence_creneau_date(int(creneau_id), date)

        remarques_value = presence_existante.get('remarques', '') if presence_existante else ''
        statut_present = 'selected' if presence_existante and presence_existante.get('statut') == 'present' else ''
        statut_absent = 'selected' if presence_existante and presence_existante.get('statut') == 'absent' else ''
        statut_retard = 'selected' if presence_existante and presence_existante.get('statut') == 'retard' else ''
        bouton_text = 'Modifier' if modification else 'Marquer'
        alert_text = 'modifiée' if modification else 'marquée'

        # Générer le HTML du formulaire
        html = f"""
        <form id="formPresence" onsubmit="soumettrePresence(event)">
            <input type="hidden" name="creneau_id" value="{creneau_id}">
            <input type="hidden" name="enseignant_id" value="{enseignant_id}">
            <input type="hidden" name="date_cours" value="{date}">

            <div class="mb-4">
                <h4 class="text-lg font-semibold text-gray-800">
                    {creneau['nom_cours']} - {creneau['jour']} {creneau['heure_debut']}-{creneau['heure_fin']}
                </h4>
                <p class="text-gray-600">Salle: {creneau['nom_salle']}</p>
            </div>

            <div class="mb-4">
                <label class="block text-sm font-medium text-gray-700 mb-2">Statut</label>
                <select name="statut" class="w-full px-4 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-uist-bleu" required>
                    <option value="present" {statut_present}>Présent</option>
                    <option value="absent" {statut_absent}>Absent</option>
                    <option value="retard" {statut_retard}>En retard</option>
                </select>
            </div>

            <div class="mb-4">
                <label class="block text-sm font-medium text-gray-700 mb-2">Remarques (optionnel)</label>
                <textarea name="remarques" class="w-full px-4 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-uist-bleu" rows="3">{remarques_value}</textarea>
            </div>

            <div class="flex justify-end space-x-3">
                <button type="button" onclick="fermerModalPresence()" class="px-4 py-2 bg-gray-300 text-gray-700 rounded-lg hover:bg-gray-400">
                    Annuler
                </button>
                <button type="submit" class="px-4 py-2 bg-uist-bleu text-white rounded-lg hover:bg-blue-600">
                    {bouton_text} la présence
                </button>
            </div>
        </form>

        <script>
        function soumettrePresence(event) {{
            event.preventDefault();
            const formData = new FormData(event.target);
            const data = Object.fromEntries(formData);

            fetch('/api/presences/marquer', {{
                method: 'POST',
                headers: {{
                    'Content-Type': 'application/json',
                }},
                body: JSON.stringify(data)
            }})
            .then(response => response.json())
            .then(result => {{
                if (result.success) {{
                    fermerModalPresence();
                    chargerPresences();
                    alert('Présence {alert_text} avec succès');
                }} else {{
                    alert('Erreur: ' + result.error);
                }}
            }})
            .catch(error => {{
                console.error('Erreur:', error);
                alert('Erreur lors de la soumission');
            }});
        }}
        </script>
        """

        return html, 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api_bp.route('/presences/marquer', methods=['POST'])
@role_requis('administration', 'sous_admin', 'ADMIN', 'SUPER_ADMIN')
def marquer_presence_api():
    """
    Marque ou modifie une présence via API
    """
    try:
        data = request.get_json()

        required_fields = ['creneau_id', 'enseignant_id', 'statut', 'date_cours']
        for field in required_fields:
            if field not in data:
                return jsonify({'success': False, 'error': f'Champ {field} requis'}), 400

        creneau_id = int(data['creneau_id'])
        enseignant_id = int(data['enseignant_id'])
        statut = data['statut']
        date_cours = data['date_cours']
        remarques = data.get('remarques', '')

        # Validation du statut
        if statut not in ['present', 'absent', 'retard']:
            return jsonify({'success': False, 'error': 'Statut invalide'}), 400

        # Marquer la présence
        presence_id = Presence.marquer_presence_api(
            creneau_id, enseignant_id, statut, date_cours, remarques, session['utilisateur_id']
        )

        if presence_id:
            # Mettre à jour les heures de l'enseignant
            from app.models import Enseignant
            Enseignant.mettre_a_jour_heures(enseignant_id)

            return jsonify({
                'success': True,
                'presence_id': presence_id,
                'message': 'Présence marquée avec succès'
            }), 200
        else:
            return jsonify({'success': False, 'error': 'Erreur lors du marquage'}), 500

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
